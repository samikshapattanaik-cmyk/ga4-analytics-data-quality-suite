#!/usr/bin/env python3
"""
validate_schema.py — GA4 Schema & Tracking Auditor core engine.

Compares an OBSERVED GA4 data export (BigQuery export, GA4 Data API pull, GTM
Preview JSON, flat gtag.js-style event capture, or a CSV/Excel/Google-Sheets
export) against a TRACKING PLAN (either a user-supplied spec, or the bundled
default in references/ga4-default-spec.json) and emits structured JSON findings
tiered into CRITICAL / WARNING / NOTICE.

This script never invents data. If a file can't be parsed, or a required
dependency is missing, it stops and prints a clear, human-readable error to
stderr along with a suggested fix, rather than guessing at rows that were
never actually there.

Dependencies: JSON, JSONL, CSV, and TSV (which covers every Google Sheets export,
since Sheets always downloads as CSV/TSV) are parsed with the Python standard
library only — no install required. pandas + openpyxl are only imported, and only
required, when the input is a genuine binary Excel file (.xlsx/.xls).

Usage:
    python validate_schema.py --observed export.json [--tracking-plan plan.csv] [--output findings.json]
    python validate_schema.py --observed export.csv --format tabular
    python validate_schema.py --observed export.jsonl --tracking-plan plan.json --output findings.json
    python validate_schema.py --observed export.json --output findings.json --excel-output findings.xlsx

Output JSON shape:
    {
      "summary": {...},
      "findings": [...],       # granular, one entry per (event, scope, parameter, issue type),
                                # deduplicated across occurrences with an occurrences_affected count.
                                # This is the full-detail view for programmatic consumers.
      "event_reports": [...]   # one entry per distinct (event, issue-pattern) combination, with
                                # every issue for that pattern grouped together and a single merged
                                # dataLayer.push fix block. This is the view meant for a human report.
    }

Exit codes:
    0 = ran successfully (findings may or may not be empty)
    1 = could not run (bad input, missing dependency, bad tracking plan, etc.)
"""

import argparse
import csv
import json
import os
import re
import sys
from collections import OrderedDict

# pandas + openpyxl are ONLY needed for genuine binary Excel files (.xlsx/.xls) —
# that's a zipped XML container, so there's no reasonable stdlib-only way to read it.
# CSV, TSV, JSON, and JSONL (including Google Sheets exports, which are always
# CSV/TSV under the hood) are handled with Python's standard library below, so the
# common path never requires installing anything.
try:
    import pandas as pd
except ImportError:  # pragma: no cover - exercised only when pandas truly absent
    pd = None


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

CRITICAL = "CRITICAL"
WARNING = "WARNING"
NOTICE = "NOTICE"
_TIER_RANK = {CRITICAL: 0, WARNING: 1, NOTICE: 2}

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
SKILL_ROOT = os.path.dirname(SCRIPT_DIR)
DEFAULT_SPEC_PATH = os.path.join(SKILL_ROOT, "references", "ga4-default-spec.json")


# ---------------------------------------------------------------------------
# Error handling — never hallucinate, always fail loud and clear
# ---------------------------------------------------------------------------

class AuditorError(Exception):
    """Raised for any condition that should stop execution with a clear message."""

    def __init__(self, message, hint=None):
        super().__init__(message)
        self.message = message
        self.hint = hint


def fail(message, hint=None):
    raise AuditorError(message, hint)


def emit_fatal_and_exit(err: AuditorError):
    sys.stderr.write(f"ERROR: {err.message}\n")
    if err.hint:
        sys.stderr.write(f"HINT: {err.hint}\n")
    sys.exit(1)


# ---------------------------------------------------------------------------
# String / casing helpers
# ---------------------------------------------------------------------------

def normalize_key(name):
    """Case- and separator-insensitive key for matching, e.g. 'Page-Location' -> 'pagelocation'."""
    if name is None:
        return ""
    return re.sub(r"[^a-z0-9]", "", str(name).lower())


def is_snake_case(name):
    if not name or not isinstance(name, str):
        return False
    return bool(re.fullmatch(r"[a-z][a-z0-9_]*", name))


def to_snake_case(name):
    """Best-effort camelCase/PascalCase/kebab-case/space-case -> snake_case."""
    s = str(name).strip()
    s = re.sub(r"[\s\-]+", "_", s)
    s = re.sub(r"(?<!^)(?<![_A-Z])(?=[A-Z])", "_", s)
    s = re.sub(r"_+", "_", s)
    return s.lower().strip("_")


def python_type_name(value):
    if isinstance(value, bool):
        return "bool"
    if isinstance(value, int):
        return "int"
    if isinstance(value, float):
        return "float"
    if value is None:
        return "null"
    if isinstance(value, str):
        # A numeric-looking string is exactly the "value passed as string" bug
        # we want to catch, so classify it specially rather than as plain "string".
        if re.fullmatch(r"-?\d+", value):
            return "numeric_string_int"
        if re.fullmatch(r"-?\d+\.\d+", value):
            return "numeric_string_float"
        if value.lower() in ("true", "false"):
            return "numeric_string_bool"
        return "string"
    return type(value).__name__


def types_compatible(expected_type, observed_value):
    """Return (is_match: bool, is_stringified_number: bool)."""
    if expected_type in (None, "any"):
        return True, False
    actual = python_type_name(observed_value)
    if expected_type == "string":
        return actual in ("string",), False
    if expected_type == "int":
        if actual == "int":
            return True, False
        if actual == "numeric_string_int":
            return False, True
        return False, False
    if expected_type == "float":
        if actual in ("int", "float"):
            return True, False
        if actual in ("numeric_string_int", "numeric_string_float"):
            return False, True
        return False, False
    if expected_type == "bool":
        if actual == "bool":
            return True, False
        if actual == "numeric_string_bool":
            return False, True
        return False, False
    return True, False


# ---------------------------------------------------------------------------
# Tracking-plan tier/required inference — the 3-layer fallback
# ---------------------------------------------------------------------------
#
# A custom tracking plan can specify tier/required explicitly, but most
# real-world tracking plans (marketing-authored Sheets/Excel docs) don't have
# those columns at all — they just have a free-text description/notes column.
# Rather than defaulting every unspecified parameter to "WARNING, required"
# (which floods the report with noise for perfectly legitimate optional
# fields like `coupon` or `affiliation`), tier/required are resolved through
# three layers, in priority order:
#
#   1. Explicit `tier`/`required` columns in the plan itself, if present.
#   2. A matching event+parameter entry in the bundled default GA4 spec,
#      borrowing its tier/required — this is real curated data, not a guess.
#   3. Cautious keyword inference from the notes/description column, as a
#      last resort for parameters with no match anywhere else.
#
# Every parameter is tagged with a `tier_source` string recording exactly
# which layer(s) supplied its tier/required, so the least-trustworthy layer
# (3) is always inspectable rather than silently indistinguishable from an
# explicit column value.

_NEGATION_OR_OPTIONAL_PATTERN = re.compile(
    r"\boptional\b|\bnot\s+required\b|\bnot\s+mandatory\b|\bisn.t\s+required\b|\bno\s+longer\s+required\b",
    re.IGNORECASE,
)
_REQUIRED_KEYWORD_PATTERN = re.compile(r"\b(required|mandatory)\b", re.IGNORECASE)
_CRITICAL_SEVERITY_KEYWORD_PATTERN = re.compile(
    r"\b(critical|high[\s-]impact|breaks?\s+(reporting|core|revenue)|revenue\s+impact|duplicate\s+revenue)\b",
    re.IGNORECASE,
)


def _infer_required_and_tier_from_notes(notes):
    """
    Best-effort layer 3 of the tracking-plan fallback. Returns (required, tier),
    each either a resolved value or None if nothing recognizable was found.
    Deliberately conservative: only acts on unambiguous keywords, and checks for
    negation/"optional" first so a note like "Not required for guest checkout"
    doesn't get misread as a required-field signal.
    """
    if not notes:
        return None, None
    text = str(notes)
    if _NEGATION_OR_OPTIONAL_PATTERN.search(text):
        return False, None
    if _REQUIRED_KEYWORD_PATTERN.search(text):
        tier = CRITICAL if _CRITICAL_SEVERITY_KEYWORD_PATTERN.search(text) else None
        return True, tier
    return None, None


def _resolve_spec_fallbacks(spec, default_spec):
    """
    Fill in tier/required for every parameter in a custom, long-format tracking
    plan (built from CSV/TSV/Excel or a flat JSON record list) using the 3-layer
    fallback described above. Mutates and returns `spec`.

    Not applied to the native `{"events": {...}}` JSON shape — choosing that
    format is itself a signal the author is writing the fully-specified schema
    directly and doesn't need inference layered on top.
    """
    default_index = {normalize_key(k): v for k, v in default_spec.items()}

    for event_name, entry in spec.items():
        default_entry = default_index.get(normalize_key(event_name))
        for bucket in ("params", "items"):
            default_bucket = (default_entry or {}).get(bucket, {})
            default_bucket_index = {normalize_key(k): v for k, v in default_bucket.items()}

            for param_name, rules in entry[bucket].items():
                explicit_tier = rules.pop("_explicit_tier", None)
                explicit_required = rules.pop("_explicit_required", None)
                notes = rules.get("notes", "")

                tier = explicit_tier
                required = explicit_required
                sources = []
                if explicit_tier is not None:
                    sources.append("tier:explicit")
                if explicit_required is not None:
                    sources.append("required:explicit")

                if tier is None or required is None:
                    default_rules = default_bucket_index.get(normalize_key(param_name))
                    if default_rules:
                        if tier is None:
                            tier = default_rules.get("tier")
                            sources.append("tier:default_spec")
                        if required is None:
                            required = default_rules.get("required", True)
                            sources.append("required:default_spec")

                if tier is None or required is None:
                    inferred_required, inferred_tier = _infer_required_and_tier_from_notes(notes)
                    if required is None and inferred_required is not None:
                        required = inferred_required
                        sources.append("required:notes_inferred")
                    if tier is None and inferred_tier is not None:
                        tier = inferred_tier
                        sources.append("tier:notes_inferred")

                if tier is None:
                    tier = WARNING
                    sources.append("tier:default_fallback")
                if required is None:
                    required = True
                    sources.append("required:default_fallback")

                rules["tier"] = tier
                rules["required"] = required
                rules["tier_source"] = "; ".join(sources)
    return spec


# ---------------------------------------------------------------------------
# Tracking plan loading
# ---------------------------------------------------------------------------

def load_default_spec():
    try:
        with open(DEFAULT_SPEC_PATH, "r", encoding="utf-8") as f:
            raw = json.load(f)
    except FileNotFoundError:
        fail(
            f"Could not find the bundled default GA4 spec at '{DEFAULT_SPEC_PATH}'.",
            "This file should ship inside references/ alongside this script. "
            "If it's missing, re-download the skill, or pass --tracking-plan to supply your own spec.",
        )
    except json.JSONDecodeError as e:
        fail(f"The bundled default spec at '{DEFAULT_SPEC_PATH}' is not valid JSON: {e}")
    if "events" not in raw:
        fail(
            f"The bundled default spec at '{DEFAULT_SPEC_PATH}' is missing its top-level 'events' key.",
            "This file ships with the skill and shouldn't normally be hand-edited into a broken state — "
            "if it has been, restore it from the skill source, or pass --tracking-plan to bypass it entirely.",
        )
    return raw["events"]


def _is_blank(value):
    return value is None or (isinstance(value, str) and value.strip() == "")


def _coerce_from_text(value):
    """
    CSV/TSV has no native types — every cell is text. pandas' read_csv used to paper
    over this by auto-inferring numeric dtypes, so a genuinely-numeric column like
    `value` would come back as a real float rather than the string "49.99". Since we
    no longer require pandas for plain-text formats, replicate that same inference
    here: numeric-looking text becomes a real int/float/bool, everything else stays
    a string. This keeps CSV/TSV behavior consistent with Excel (where openpyxl already
    hands back typed cells) and avoids flagging every ordinary numeric CSV column as a
    false "value sent as string" type mismatch.
    """
    if value is None:
        return None
    s = str(value).strip()
    if s == "":
        return None
    if re.fullmatch(r"-?\d+", s):
        return int(s)
    if re.fullmatch(r"-?\d+\.\d+", s):
        return float(s)
    if s.lower() in ("true", "false"):
        return s.lower() == "true"
    return s


def _read_delimited_rows(path, delimiter):
    """Read a CSV/TSV file into a list of dicts using only the standard library."""
    try:
        with open(path, "r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f, delimiter=delimiter)
            if reader.fieldnames is None:
                fail(f"'{path}' appears to be empty (no header row found).")
            fieldnames = list(reader.fieldnames)
            rows = [{k: _coerce_from_text(v) for k, v in raw_row.items()} for raw_row in reader]
    except UnicodeDecodeError as e:
        fail(
            f"Could not read '{path}' as UTF-8 text: {e}",
            "This might actually be a binary Excel file saved with a .csv extension — "
            "try re-exporting it as a genuine .xlsx, or re-save it as CSV from your spreadsheet app.",
        )
    return rows, fieldnames


def _read_excel_rows(path):
    """Read a genuine .xlsx/.xls file into a list of dicts. Requires pandas + openpyxl."""
    if pd is None:
        fail(
            "This is a binary Excel file, which needs pandas + openpyxl to read (there's no "
            "standard-library way to parse the .xlsx/.xls container format).",
            "Install them in your environment (e.g. `pip install pandas openpyxl`, using a "
            "virtualenv or `--break-system-packages` if your environment requires it), "
            "or — usually simpler — re-export this file as CSV from Excel/Google Sheets and "
            "point me at that instead; CSV needs no extra dependencies at all.",
        )
    try:
        df = pd.read_excel(path)
    except Exception as e:
        fail(
            f"Could not read '{path}' as Excel: {e}",
            "Is openpyxl installed? Or is this actually a renamed CSV/TSV in disguise? "
            "Try re-exporting the file and running again.",
        )
    df.columns = [str(c).strip() for c in df.columns]
    records = df.to_dict(orient="records")
    return records, list(df.columns)


def _clean_type(value):
    v = str(value).strip().lower() if value not in (None, "") else "any"
    return v if v in ("string", "int", "float", "bool", "any") else "any"


def _parse_explicit_tier(value):
    """Returns None (unset) rather than a default, so the fallback chain knows to keep looking."""
    if value in (None, ""):
        return None
    v = str(value).strip().upper()
    return v if v in (CRITICAL, WARNING, NOTICE) else None


def _parse_explicit_required(value):
    """Returns None (unset) rather than a default, so the fallback chain knows to keep looking."""
    if value in (None, ""):
        return None
    if isinstance(value, bool):
        return value
    s = str(value).strip().lower()
    if s in ("false", "0", "no", "n", "optional"):
        return False
    if s in ("true", "1", "yes", "y", "required"):
        return True
    return None


def _spec_from_long_records(records):
    """
    records: iterable of dicts with at least event_name + param_name.
    Optional: scope (event|item), tier, data_type, notes, required.

    tier/required are kept possibly-None here (not yet defaulted) — the 3-layer
    fallback in _resolve_spec_fallbacks fills them in afterward.
    """
    spec = OrderedDict()
    for row in records:
        event = (row.get("event_name") or row.get("event") or "").strip()
        if not event:
            continue
        param = (row.get("param_name") or row.get("parameter") or row.get("param") or "").strip()
        if not param:
            continue
        scope = str(row.get("scope") or "event").strip().lower()
        explicit_tier = _parse_explicit_tier(row.get("tier"))
        dtype = _clean_type(row.get("data_type") or row.get("type"))
        notes = str(row.get("notes") or row.get("description") or row.get("description / when it fires") or "").strip()
        explicit_required = _parse_explicit_required(row.get("required"))

        spec.setdefault(event, {"category": "custom", "params": {}, "items": {}})
        bucket = "items" if scope.startswith("item") else "params"
        spec[event][bucket][param] = {
            "type": dtype,
            "notes": notes,
            "_explicit_tier": explicit_tier,
            "_explicit_required": explicit_required,
        }
    if not spec:
        fail(
            "The tracking plan was read but produced zero usable rows.",
            "Double check it has an 'event_name' column/key and a 'param_name' column/key with values in them.",
        )
    return spec


def load_custom_tracking_plan(path):
    if not os.path.exists(path):
        fail(f"Tracking plan file not found: '{path}'.")

    ext = os.path.splitext(path)[1].lower()

    if ext == ".json":
        try:
            with open(path, "r", encoding="utf-8") as f:
                raw = json.load(f)
        except json.JSONDecodeError as e:
            fail(f"Could not parse '{path}' as JSON: {e}",
                 "Check for a trailing comma or unclosed bracket, or export the plan as CSV instead.")
        # Two shapes accepted: {"events": {...}} matching our native spec shape
        # (taken as fully-specified, no fallback inference applied), or a flat
        # list of long-format records (fallback inference applied, same as CSV/Excel).
        if isinstance(raw, dict) and "events" in raw:
            return raw["events"]
        if isinstance(raw, list):
            spec = _spec_from_long_records(raw)
            return _resolve_spec_fallbacks(spec, load_default_spec())
        fail(
            f"'{path}' is valid JSON but not in a recognized tracking-plan shape.",
            'Expected either {"events": {...}} (native spec format) or a flat list of '
            '{"event_name": ..., "param_name": ..., "scope": ..., "tier": ..., "data_type": ..., '
            '"required": ...} records.',
        )

    if ext in (".csv", ".tsv", ".xlsx", ".xls"):
        if ext == ".csv":
            records, columns = _read_delimited_rows(path, delimiter=",")
        elif ext == ".tsv":
            records, columns = _read_delimited_rows(path, delimiter="\t")
        else:
            records, columns = _read_excel_rows(path)

        columns = [str(c).strip() for c in columns]
        cols_lower = {c.lower(): c for c in columns}
        has_param_col = any(
            k in cols_lower for k in ("param_name", "parameter", "param", "parameter_name")
        )
        if "event_name" not in cols_lower or not has_param_col:
            fail(
                f"Tracking plan '{path}' is missing required column(s).",
                "Expected at least an 'event_name' column and a 'param_name' "
                "(or 'parameter'/'param'/'parameter_name') column. Optional columns: "
                "'scope'/'parameter_scope' (event|item), 'tier' (CRITICAL|WARNING|NOTICE), "
                "'data_type'/'parameter_type' (string|int|float|bool), 'notes'/'description', "
                "'required' (true|false). Any of tier/required/notes may be omitted entirely — "
                "missing tier/required are inferred from the bundled default GA4 spec and, failing "
                "that, cautiously from the notes/description text. "
                f"Columns found: {columns}",
            )
        # Normalize keys to the lowercase names _spec_from_long_records expects,
        # including a couple of real-world column-name variants (e.g. a plan
        # authored with 'parameter_name'/'parameter_scope'/'parameter_type'
        # instead of 'param_name'/'scope'/'data_type').
        normalized = []
        for row in records:
            norm_row = {str(k).strip().lower(): v for k, v in row.items()}
            if "param_name" not in norm_row and "parameter_name" in norm_row:
                norm_row["param_name"] = norm_row["parameter_name"]
            if "scope" not in norm_row and "parameter_scope" in norm_row:
                norm_row["scope"] = norm_row["parameter_scope"]
            if "data_type" not in norm_row and "parameter_type" in norm_row:
                norm_row["data_type"] = norm_row["parameter_type"]
            normalized.append(norm_row)
        spec = _spec_from_long_records(normalized)
        return _resolve_spec_fallbacks(spec, load_default_spec())

    fail(
        f"Unrecognized tracking plan file type '{ext}'.",
        "Supported formats: .json, .csv, .tsv, .xlsx, .xls. "
        "If this came from Google Sheets, use File > Download > CSV first.",
    )


# ---------------------------------------------------------------------------
# Observed data loading — format auto-detection
# ---------------------------------------------------------------------------

def detect_and_load_observed(path, format_override=None):
    if not os.path.exists(path):
        fail(f"Observed data file not found: '{path}'.")

    ext = os.path.splitext(path)[1].lower()
    fmt = format_override

    if fmt is None:
        if ext in (".json", ".jsonl", ".ndjson"):
            fmt = _sniff_json_shape(path, ext)
        elif ext in (".csv", ".tsv", ".xlsx", ".xls"):
            fmt = "tabular"
        else:
            fail(
                f"Unrecognized observed-data file type '{ext}'.",
                "Supported formats: BigQuery JSON/JSONL export (.json/.jsonl), GTM Preview JSON (.json), "
                "GA4 Data API JSON pull (.json), flat gtag.js-style event capture (.json), or a "
                "CSV/TSV/Excel/Google Sheets export (.csv/.tsv/.xlsx/.xls). "
                "Pass --format to force one explicitly if auto-detection guesses wrong.",
            )

    if fmt == "bq_json":
        return _load_bq_json(path, ext), fmt
    if fmt == "ga4_api_json":
        return _load_ga4_api_json(path), fmt
    if fmt == "gtm_preview_json":
        return _load_gtm_preview_json(path), fmt
    if fmt == "flat_json":
        return _load_flat_json(path, ext), fmt
    if fmt == "tabular":
        return _load_tabular(path, ext), fmt
    fail(f"Unknown --format override '{fmt}'. Expected one of: bq_json, ga4_api_json, gtm_preview_json, flat_json, tabular.")


def _read_json_or_jsonl(path, ext):
    try:
        with open(path, "r", encoding="utf-8") as f:
            text = f.read().strip()
    except UnicodeDecodeError as e:
        fail(f"Could not read '{path}' as UTF-8 text: {e}",
             "Is this actually a binary/Excel file with a .json extension? Try re-exporting it.")
    if not text:
        fail(f"'{path}' is empty.")
    if ext in (".jsonl", ".ndjson") or (text[0] != "[" and text[0] != "{" and "\n" in text):
        rows = []
        for i, line in enumerate(text.splitlines(), start=1):
            line = line.strip()
            if not line:
                continue
            try:
                rows.append(json.loads(line))
            except json.JSONDecodeError as e:
                fail(f"Line {i} of '{path}' is not valid JSON: {e}",
                     "This looked like newline-delimited JSON (JSONL) — check that every line is one complete JSON object.")
        return rows
    try:
        return json.loads(text)
    except json.JSONDecodeError as e:
        fail(f"Could not parse '{path}' as JSON: {e}",
             "If this is a newline-delimited BigQuery export, try renaming it to .jsonl and re-running.")


def _sniff_json_shape(path, ext):
    parsed = _read_json_or_jsonl(path, ext)
    sample = None
    if isinstance(parsed, list) and parsed:
        sample = parsed[0]
    elif isinstance(parsed, dict):
        if "rows" in parsed and ("dimensionHeaders" in parsed or "metricHeaders" in parsed):
            return "ga4_api_json"
        if "dataLayer" in parsed and isinstance(parsed["dataLayer"], list) and parsed["dataLayer"]:
            sample = parsed["dataLayer"][0]
        else:
            sample = parsed

    if isinstance(sample, dict):
        # Order matters: check the most structurally-specific shapes first.
        if "event_params" in sample:
            # Genuine BigQuery export rows always carry an event_params array
            # (even if empty) — that's the reliable signal for this shape,
            # not just the presence of "event_name".
            return "bq_json"
        if "event" in sample and ("ecommerce" in sample or "gtm.uniqueEventId" in sample or "gtm.start" in str(sample)):
            return "gtm_preview_json"
        if "event" in sample:
            # A flat dataLayer-style push: {"event": "add_to_cart", ...flat params...}
            return "gtm_preview_json"
        if "event_name" in sample:
            # Flat shape with no event_params/ecommerce wrapper at all — params and
            # items sit directly on the row. This is the natural output of many
            # lightweight capture tools (e.g. a browser extension or script that
            # intercepts gtag('event', name, paramsObj) calls and logs them as-is),
            # and is NOT the same as a real BigQuery export just because it also
            # happens to have an "event_name" key.
            return "flat_json"

    fail(
        f"Could not confidently auto-detect the JSON structure of '{path}'.",
        "Pass --format explicitly: bq_json (BigQuery export rows with event_params/items), "
        "gtm_preview_json (GTM Preview / dataLayer dump), ga4_api_json (GA4 Data API pull with "
        "rows/dimensionHeaders), or flat_json (flat {event_name, ...params, items} objects).",
    )


# --- BigQuery-style export -------------------------------------------------

def _bq_param_value(value_struct):
    if not isinstance(value_struct, dict):
        return value_struct
    for key in ("string_value", "int_value", "float_value", "double_value"):
        if key in value_struct and value_struct[key] is not None:
            val = value_struct[key]
            if key == "int_value":
                try:
                    return int(val)
                except (TypeError, ValueError):
                    return val
            if key in ("float_value", "double_value"):
                try:
                    return float(val)
                except (TypeError, ValueError):
                    return val
            return val
    return None


def _load_bq_json(path, ext):
    rows = _read_json_or_jsonl(path, ext)
    if isinstance(rows, dict):
        rows = rows.get("rows") or rows.get("data") or [rows]
    if not isinstance(rows, list):
        fail(f"Expected '{path}' to contain a list of event rows for BigQuery-style export.")

    events = []
    for i, row in enumerate(rows):
        if not isinstance(row, dict):
            continue
        event_name = row.get("event_name") or row.get("eventName")
        if not event_name:
            continue
        params = {}
        for p in row.get("event_params") or []:
            key = p.get("key")
            if not key:
                continue
            params[key] = _bq_param_value(p.get("value"))

        items = []
        for item in row.get("items") or []:
            if isinstance(item, dict):
                items.append(dict(item))

        events.append({
            "event_name": event_name,
            "source_ref": f"row {i + 1}",
            "params": params,
            "items": items,
        })

    if not events:
        fail(
            f"'{path}' parsed as JSON but no rows contained a usable 'event_name'.",
            "Confirm this is really a GA4 BigQuery export table dump (rows with event_name/event_params/items).",
        )
    return events


# --- GTM Preview / dataLayer dump -----------------------------------------

def _flatten_gtm_object(obj, prefix=""):
    flat = {}
    items = []
    for k, v in obj.items():
        if k in ("event", "gtm.uniqueEventId", "gtm.start", "eventCallback", "eventTimeout"):
            continue
        if k == "ecommerce" and isinstance(v, dict):
            for ek, ev in v.items():
                if ek == "items" and isinstance(ev, list):
                    for item in ev:
                        if isinstance(item, dict):
                            items.append(dict(item))
                else:
                    flat[ek] = ev
            continue
        if k == "items" and isinstance(v, list):
            for item in v:
                if isinstance(item, dict):
                    items.append(dict(item))
            continue
        if isinstance(v, (dict, list)):
            continue  # skip other nested structures we don't have a defined shape for
        flat[f"{prefix}{k}"] = v
    return flat, items


def _load_gtm_preview_json(path):
    parsed = _read_json_or_jsonl(path, os.path.splitext(path)[1].lower())
    if isinstance(parsed, dict) and "dataLayer" in parsed:
        pushes = parsed["dataLayer"]
    elif isinstance(parsed, list):
        pushes = parsed
    else:
        pushes = [parsed]

    events = []
    for i, push in enumerate(pushes):
        if not isinstance(push, dict):
            continue
        event_name = push.get("event")
        if not event_name:
            continue
        params, items = _flatten_gtm_object(push)
        events.append({
            "event_name": event_name,
            "source_ref": f"dataLayer push #{i + 1}",
            "params": params,
            "items": items,
        })

    if not events:
        fail(
            f"'{path}' parsed as JSON but no entries contained an 'event' key.",
            "Confirm this is a GTM Preview / dataLayer export where each push has an 'event' field.",
        )
    return events


# --- Flat gtag.js-style event capture --------------------------------------

def _load_flat_json(path, ext):
    """
    Flat JSON shape: each row is {"event_name": ..., <param>: value, ..., "items": [...]}
    with no event_params/ecommerce wrapper — params and items sit directly on the
    row. Common output of tools that intercept and log raw gtag('event', name, params)
    calls (e.g. a QA browser extension or console override), rather than reading a
    BigQuery export or a GTM Preview panel.
    """
    rows = _read_json_or_jsonl(path, ext)
    if isinstance(rows, dict):
        rows = rows.get("rows") or rows.get("data") or rows.get("events") or [rows]
    if not isinstance(rows, list):
        fail(f"Expected '{path}' to contain a list of flat event objects.")

    events = []
    for i, row in enumerate(rows):
        if not isinstance(row, dict):
            continue
        event_name = row.get("event_name") or row.get("event")
        if not event_name:
            continue
        items = []
        params = {}
        for k, v in row.items():
            if k in ("event_name", "event"):
                continue
            if k == "items" and isinstance(v, list):
                items = [dict(it) for it in v if isinstance(it, dict)]
                continue
            if isinstance(v, (dict, list)):
                continue  # skip unexpected nested structures we don't have a defined shape for
            params[k] = v
        events.append({
            "event_name": event_name,
            "source_ref": f"row {i + 1}",
            "params": params,
            "items": items,
        })

    if not events:
        fail(
            f"'{path}' parsed as JSON but no entries contained a usable 'event_name'/'event' key.",
            "Confirm this is a flat event export where each object has event_name plus its "
            "parameters directly as keys (and optionally an items[] array).",
        )
    return events


# --- GA4 Data API JSON pull (best-effort) ----------------------------------

def _load_ga4_api_json(path):
    parsed = _read_json_or_jsonl(path, ".json")
    dim_headers = [h.get("name") for h in parsed.get("dimensionHeaders", [])]
    met_headers = [h.get("name") for h in parsed.get("metricHeaders", [])]
    rows = parsed.get("rows", [])
    if not rows:
        fail(
            f"'{path}' looks like a GA4 Data API response but has no rows.",
            "Nothing to audit — re-run the API report with a non-empty date range.",
        )

    event_name_idx = None
    for i, name in enumerate(dim_headers):
        if name and "eventname" in name.lower().replace("_", ""):
            event_name_idx = i
            break
    if event_name_idx is None:
        fail(
            f"Could not find an eventName dimension in '{path}'.",
            f"GA4 Data API pulls need an 'eventName' dimension included in the report. "
            f"Dimension headers found: {dim_headers}",
        )

    events = []
    for i, row in enumerate(rows):
        dim_values = [dv.get("value") for dv in row.get("dimensionValues", [])]
        met_values = [mv.get("value") for mv in row.get("metricValues", [])]
        event_name = dim_values[event_name_idx] if event_name_idx < len(dim_values) else None
        if not event_name:
            continue
        params = {}
        for h, v in zip(dim_headers, dim_values):
            if h and h != dim_headers[event_name_idx]:
                params[h] = v
        for h, v in zip(met_headers, met_values):
            if h:
                try:
                    params[h] = float(v)
                except (TypeError, ValueError):
                    params[h] = v
        events.append({
            "event_name": event_name,
            "source_ref": f"api row {i + 1}",
            "params": params,
            "items": [],
        })

    if not events:
        fail(f"'{path}' had rows but none resolved to a usable event name.")
    return events


# --- CSV / Excel / Google Sheets export (tabular) --------------------------

def _load_tabular(path, ext):
    """
    Load a CSV/TSV/Excel observed-data export. CSV and TSV — which covers every
    Google Sheets export, since Sheets always exports to one of those two — are read
    with the standard library and need no extra dependencies. Only genuine binary
    Excel files (.xlsx/.xls) fall through to pandas + openpyxl, since that's a zipped
    binary container with no reasonable stdlib-only reader.
    """
    if ext == ".csv":
        records, columns = _read_delimited_rows(path, delimiter=",")
    elif ext == ".tsv":
        records, columns = _read_delimited_rows(path, delimiter="\t")
    else:
        records, columns = _read_excel_rows(path)

    columns = [str(c).strip() for c in columns]
    cols_lower = {c.lower(): c for c in columns}

    event_col = next((cols_lower[c] for c in ("event_name", "eventname", "event") if c in cols_lower), None)
    if event_col is None:
        fail(
            f"Could not find an event name column in '{path}'.",
            f"Expected a column called 'event_name' (or 'eventName'/'event'). Columns found: {columns}. "
            "If this file uses a different column name, rename it and re-run, or ask me to help map columns.",
        )

    long_format_key_col = next(
        (cols_lower[c] for c in ("param_name", "parameter", "param", "param_key") if c in cols_lower), None
    )
    long_format_val_col = next((cols_lower[c] for c in ("param_value", "value") if c in cols_lower), None)

    events = []
    if long_format_key_col and long_format_val_col:
        # Long format: one row per (event occurrence, parameter). Group rows back into
        # occurrences by an explicit occurrence/event id column when present. Without
        # one, there's no way to tell apart two separate firings of the same event, so
        # rows sharing an event_name are merged into a single representative occurrence
        # for that event — accurate for a "here's one example row per parameter" style
        # export, but add an occurrence_id column if you need per-firing granularity.
        scope_col = next((cols_lower[c] for c in ("scope",) if c in cols_lower), None)
        occurrence_col = next(
            (cols_lower[c] for c in ("occurrence_id", "event_id", "row_id") if c in cols_lower), None
        )
        # An optional column to tell multiple items in the same occurrence apart. Without
        # one, all item-scoped rows for an occurrence are assumed to describe a single item
        # and get merged into one item object — add an item_index column if a real
        # occurrence has more than one item and you need them told apart.
        item_index_col = next(
            (cols_lower[c] for c in ("item_index", "item_row", "item_id_index") if c in cols_lower), None
        )
        grouped = OrderedDict()
        for row in records:
            event_name = row.get(event_col)
            if _is_blank(event_name):
                continue
            occ_key = row.get(occurrence_col) if occurrence_col else event_name
            scope = str(row.get(scope_col) or "event").strip().lower()
            param = row.get(long_format_key_col)
            val = row.get(long_format_val_col)
            if _is_blank(param):
                continue
            bucket = grouped.setdefault(
                occ_key, {"event_name": event_name, "params": {}, "items": OrderedDict()}
            )
            if scope.startswith("item"):
                item_key = row.get(item_index_col) if item_index_col else "0"
                bucket["items"].setdefault(item_key, {})[param] = val
            else:
                bucket["params"][param] = val
        for occ_key, bucket in grouped.items():
            events.append({
                "event_name": bucket["event_name"],
                "source_ref": f"row group '{occ_key}'",
                "params": bucket["params"],
                "items": list(bucket["items"].values()),
            })
    else:
        # Wide format: every other column is a parameter. A column literally named
        # 'items' containing a JSON string is treated as the items array.
        param_cols = [c for c in columns if c != event_col]
        items_col = next((c for c in param_cols if c.lower() == "items"), None)
        if items_col:
            param_cols.remove(items_col)

        for i, row in enumerate(records):
            event_name = row.get(event_col)
            if _is_blank(event_name):
                continue
            params = {}
            for c in param_cols:
                val = row.get(c)
                if _is_blank(val):
                    continue
                params[c] = val
            items = []
            raw_items = row.get(items_col) if items_col else None
            if not _is_blank(raw_items):
                try:
                    parsed_items = json.loads(raw_items)
                    if isinstance(parsed_items, list):
                        items = [it for it in parsed_items if isinstance(it, dict)]
                except (json.JSONDecodeError, TypeError):
                    pass  # leave items empty rather than guessing at malformed JSON
            events.append({
                "event_name": event_name,
                "source_ref": f"row {i + 2}",  # +2: 1-indexed plus header row
                "params": params,
                "items": items,
            })

    if not events:
        fail(f"'{path}' was read successfully but produced zero usable event rows.")
    return events


# ---------------------------------------------------------------------------
# Diff engine
# ---------------------------------------------------------------------------
#
# Runs in three passes:
#   1. Compute raw per-occurrence issues (every issue tagged with a `kind`,
#      a stable dedup `key`, and — internally only — the actual observed
#      value where relevant for building fix blocks later).
#   2. Deduplicate those across all occurrences into the granular `findings`
#      list (unchanged schema/behavior from before): one entry per distinct
#      (event, scope, parameter, issue kind), with an occurrences_affected count.
#   3. Group occurrences by (event, exact set of issues triggered) into
#      `event_reports`: one row per distinct breakage pattern, each carrying
#      every issue for that pattern plus a single merged dataLayer.push fix
#      block built from a representative real occurrence.

def build_spec_index(spec):
    """normalized event name -> (canonical_event_name, spec_entry)"""
    index = {}
    for event_name, entry in spec.items():
        index[normalize_key(event_name)] = (event_name, entry)
    return index


def _fix_snippet_missing(event_name, scope, param_name, expected_type, is_item):
    if is_item:
        return (
            f"Add `{param_name}` to each object in the `items[]` array for `{event_name}`, e.g.:\n"
            f"  items: [{{ ..., \"{param_name}\": <{expected_type}> }}]"
        )
    return (
        f"dataLayer.push({{ event: '{event_name}', '{param_name}': <{expected_type}> }});"
    )


def _fix_snippet_casing(observed_name, canonical_name, is_item):
    where = "item object" if is_item else "dataLayer push"
    return (
        f"Rename `{observed_name}` to `{canonical_name}` in the {where}. "
        f"GA4 treats these as two different parameters today, so data is currently split across both."
    )


def _fix_snippet_type(param_name, expected_type, observed_value, is_stringified_number):
    if is_stringified_number:
        return (
            f"Send `{param_name}` as a raw number, not a string: "
            f"use {repr(_try_unquote_number(observed_value))} instead of {repr(observed_value)}."
        )
    return f"Send `{param_name}` as a {expected_type}; got {python_type_name(observed_value)} ({observed_value!r})."


def _try_unquote_number(value):
    try:
        if "." in str(value):
            return float(value)
        return int(value)
    except (TypeError, ValueError):
        return value


def _diff_occurrence(occurrence, spec_index):
    """
    Pure per-occurrence diff: returns a list of issue dicts for this one
    occurrence, each with a stable `key` (for cross-occurrence dedup) and a
    `kind` (for pattern-matching in the aggregation pass). Never mutates
    anything global — this function knows nothing about other occurrences.
    """
    issues = []
    raw_event_name = occurrence["event_name"]
    norm = normalize_key(raw_event_name)
    match = spec_index.get(norm)

    if match is None:
        issues.append({
            "key": ("unrecognized_event", raw_event_name),
            "kind": "unrecognized_event",
            "event_name": raw_event_name,
            "severity": NOTICE,
            "scope": "event",
            "parameter": None,
            "issue": f"Event '{raw_event_name}' was not found in the tracking plan.",
            "expected": "A matching event definition in the tracking plan.",
            "suggested_fix": (
                "If this is an intentional custom event, add it to your tracking plan so future audits "
                "recognize it. If it's an unintended typo/variant of a known event, rename it to match."
            ),
        })
        return raw_event_name, issues

    canonical_name, spec_entry = match
    if raw_event_name != canonical_name:
        issues.append({
            "key": ("event_casing", raw_event_name, canonical_name),
            "kind": "event_casing",
            "event_name": raw_event_name,
            "severity": NOTICE,
            "scope": "event",
            "parameter": None,
            "issue": f"Event name '{raw_event_name}' drifts from the tracking plan's canonical '{canonical_name}'.",
            "expected": canonical_name,
            "suggested_fix": _fix_snippet_casing(raw_event_name, canonical_name, is_item=False),
        })

    issues.extend(
        _diff_param_scope(canonical_name, spec_entry.get("params", {}), occurrence.get("params", {}), scope="event")
    )

    item_spec = spec_entry.get("items", {})
    if item_spec:
        if occurrence.get("items"):
            for item in occurrence["items"]:
                issues.extend(_diff_param_scope(canonical_name, item_spec, item, scope="item"))
        else:
            required_item_params = {k: v for k, v in item_spec.items() if v.get("required", True)}
            if required_item_params:
                worst_tier = min((p["tier"] for p in required_item_params.values()), key=lambda t: _TIER_RANK[t])
                missing_severity = CRITICAL if worst_tier == CRITICAL else NOTICE
                worst_source = next(
                    (p.get("tier_source", "") for p in required_item_params.values() if p["tier"] == worst_tier), ""
                )
                notes_caveat = (
                    " (tier inferred from notes — verify)"
                    if missing_severity == CRITICAL and "tier:notes_inferred" in worst_source
                    else ""
                )
                issues.append({
                    "key": ("missing_items_array", canonical_name),
                    "kind": "missing_items_array",
                    "event_name": canonical_name,
                    "severity": missing_severity,
                    "scope": "item",
                    "parameter": None,
                    "issue": f"'{canonical_name}' has no items[] array at all, but the tracking plan expects item-scoped parameters.{notes_caveat}",
                    "expected": "A non-empty items[] array.",
                    "suggested_fix": (
                        f"Populate the `items[]` array on '{canonical_name}' with at least "
                        f"{', '.join(required_item_params.keys())}."
                    ),
                })

    return canonical_name, issues


# Known common misnomers for standard GA4 parameter names, mapping a normalized
# wrong name to the normalized correct one. Deliberately a curated list rather than
# generic fuzzy-matching: plain string-similarity scoring (e.g. difflib) rates
# 'item_name' vs 'item_id' at 0.62 and 'item_name' vs 'item_category' at 0.64 —
# both close to or higher than 'item_code' vs 'item_id' at 0.75 — so a similarity
# cutoff loose enough to catch real typos would just as readily misidentify one
# genuinely distinct, correctly-named GA4 field as a typo of another whenever a
# sparse custom plan simply doesn't declare it. A curated list trades recall
# (it won't catch every possible typo) for precision (it won't invent a
# misleading "did you mean" for a parameter that was actually spelled right).
_KNOWN_PARAM_ALIASES = {
    "itemcode": "item_id",
    "productid": "item_id",
    "prodid": "item_id",
    "sku": "item_id",
    "itemsku": "item_id",
    "producttitle": "item_name",
    "productname": "item_name",
    "itemtitle": "item_name",
    "qty": "quantity",
    "itemqty": "quantity",
    "itemprice": "price",
    "unitprice": "price",
    "txnid": "transaction_id",
    "orderid": "transaction_id",
    "ordernumber": "transaction_id",
    "transactionnumber": "transaction_id",
    "curr": "currency",
    "amount": "value",
    "total": "value",
    "revenue": "value",
    "pageurl": "page_location",
    "pagepath": "page_location",
    "url": "page_location",
    "searchquery": "search_term",
    "query": "search_term",
}


def _find_unrecognized_param_issues(canonical_event_name, param_spec, observed_index, matched_norm_keys, scope):
    """
    Flag observed keys that never matched any spec parameter and are a known
    common misnomer for one the plan actually declares — e.g. item_code when
    the plan expects item_id. Only fires when the corrected name is genuinely
    part of the active spec, so the suggestion is always actionable right now.
    """
    issues = []
    spec_norm_to_name = {normalize_key(p): p for p in param_spec}
    for norm_key, (observed_name, _observed_value) in observed_index.items():
        if norm_key in matched_norm_keys:
            continue
        alias_target_raw = _KNOWN_PARAM_ALIASES.get(norm_key)
        if alias_target_raw is None:
            continue
        alias_target_norm = normalize_key(alias_target_raw)
        if alias_target_norm not in spec_norm_to_name:
            continue
        canonical_target = spec_norm_to_name[alias_target_norm]
        where = "each items[] entry" if scope == "item" else "the event payload"
        issues.append({
            "key": ("unrecognized_param", canonical_event_name, scope, observed_name),
            "kind": "unrecognized_param",
            "event_name": canonical_event_name,
            "severity": NOTICE,
            "scope": scope,
            "parameter": observed_name,
            "issue": (
                f"'{observed_name}' isn't a standard GA4 parameter — did you mean '{canonical_target}'? "
                f"Custom keys inside {'items[]' if scope == 'item' else 'the event payload'} that don't match "
                f"a real GA4 field are silently ignored by standard reports."
            ),
            "expected": f"'{canonical_target}' (closest recognized parameter)",
            "suggested_fix": f"Rename `{observed_name}` to `{canonical_target}` in {where} — GA4 only recognizes the standard field name.",
        })
    return issues


def _diff_param_scope(canonical_event_name, param_spec, observed_params, scope):
    """Pure per-occurrence, per-scope diff. Returns a list of issue dicts."""
    issues = []
    observed_index = {normalize_key(k): (k, v) for k, v in (observed_params or {}).items()}
    matched_norm_keys = set()

    for spec_param, rules in param_spec.items():
        norm_param = normalize_key(spec_param)
        found = observed_index.get(norm_param)
        if found is not None:
            matched_norm_keys.add(norm_param)

        if found is None:
            # required defaults to True when unset — a param with no explicit "required"
            # is treated as expected-by-default. Only an explicit required: false skips
            # the "missing" finding entirely, since some params (e.g. quantity on
            # view_item) are legitimately optional and shouldn't nag on every audit.
            if not rules.get("required", True):
                continue
            # Missing a CRITICAL-tier (truly load-bearing) param stays CRITICAL.
            # Missing anything else (WARNING/NOTICE-tier — i.e. "recommended but
            # not load-bearing") is downgraded to NOTICE: WARNING is reserved for
            # something actually being wrong (a bad type), not merely absent.
            missing_severity = CRITICAL if rules["tier"] == CRITICAL else NOTICE
            tier_source = rules.get("tier_source", "")
            notes_caveat = (
                " (tier inferred from notes — verify)"
                if missing_severity == CRITICAL and "tier:notes_inferred" in tier_source
                else ""
            )
            issues.append({
                "key": ("missing_param", canonical_event_name, scope, spec_param),
                "kind": "missing_param",
                "event_name": canonical_event_name,
                "severity": missing_severity,
                "scope": scope,
                "parameter": spec_param,
                "issue": f"Missing {'item-scoped' if scope == 'item' else 'event-scoped'} parameter '{spec_param}'.{notes_caveat}",
                "expected": f"'{spec_param}' present ({rules['type']}).",
                "suggested_fix": _fix_snippet_missing(canonical_event_name, scope, spec_param, rules["type"], scope == "item"),
            })
            continue

        observed_name, observed_value = found
        if observed_name != spec_param:
            issues.append({
                "key": ("param_casing", canonical_event_name, scope, spec_param, observed_name),
                "kind": "param_casing",
                "event_name": canonical_event_name,
                "severity": NOTICE,
                "scope": scope,
                "parameter": spec_param,
                "issue": f"Parameter '{observed_name}' drifts from the tracking plan's canonical '{spec_param}'.",
                "expected": spec_param,
                "suggested_fix": _fix_snippet_casing(observed_name, spec_param, is_item=(scope == "item")),
            })

        is_match, is_stringified_number = types_compatible(rules["type"], observed_value)
        if not is_match:
            issues.append({
                "key": ("type_mismatch", canonical_event_name, scope, spec_param),
                "kind": "type_mismatch",
                "event_name": canonical_event_name,
                "severity": WARNING,
                "scope": scope,
                "parameter": spec_param,
                # The actual bad value is folded directly into the issue text (rather than a
                # separate "Observed" field) since it's the one piece of diagnostic detail
                # that isn't already implied by "Issue"/"Expected" for every other issue kind.
                "issue": f"'{spec_param}' expected type {rules['type']}, got {python_type_name(observed_value)} ({observed_value!r}).",
                "expected": rules["type"],
                "suggested_fix": _fix_snippet_type(spec_param, rules["type"], observed_value, is_stringified_number),
                "_observed_value": observed_value,  # internal only, used to build merged fix blocks; stripped before output
            })

    issues.extend(
        _find_unrecognized_param_issues(canonical_event_name, param_spec, observed_index, matched_norm_keys, scope)
    )
    return issues


def _dedup_global_findings(occurrence_issues):
    """Pass 2: the granular, deduplicated findings list (existing schema)."""
    findings = OrderedDict()
    for _occurrence, _canonical_name, issues in occurrence_issues:
        for issue in issues:
            key = issue["key"]
            if key not in findings:
                findings[key] = {
                    "event_name": issue["event_name"],
                    "severity": issue["severity"],
                    "scope": issue["scope"],
                    "parameter": issue["parameter"],
                    "issue": issue["issue"],
                    "expected": issue["expected"],
                    "suggested_fix": issue["suggested_fix"],
                    "occurrences_affected": 0,
                }
            findings[key]["occurrences_affected"] += 1
    findings_list = list(findings.values())
    findings_list.sort(key=lambda f: (_TIER_RANK[f["severity"]], f["event_name"], f.get("parameter") or ""))
    return findings_list


def _placeholder_for_type(type_name):
    return f"<{type_name}>"


def _js_literal(value):
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, (int, float)):
        return repr(value)
    if value is None:
        return "null"
    return json.dumps(str(value))


def _merge_scope_fields(param_spec, observed_dict, missing_params, type_bad_params):
    """
    Build one merged field dict for a single scope (event or one item), combining:
    valid observed values as-is, missing fields as typed placeholders, and
    present-but-wrong-typed fields replaced with a corrected placeholder.
    Genuinely optional-and-absent fields are simply left out.
    """
    observed_index = {normalize_key(k): (k, v) for k, v in (observed_dict or {}).items()}
    merged = OrderedDict()
    for spec_param, rules in param_spec.items():
        norm_param = normalize_key(spec_param)
        found = observed_index.get(norm_param)
        if spec_param in type_bad_params:
            merged[spec_param] = _placeholder_for_type(rules["type"])
        elif spec_param in missing_params:
            if not rules.get("required", True):
                continue
            merged[spec_param] = _placeholder_for_type(rules["type"])
        elif found is not None:
            merged[spec_param] = found[1]
    return merged


def _render_field_entries(fields):
    rendered = []
    for k, v in fields.items():
        if isinstance(v, str) and v.startswith("<") and v.endswith(">"):
            rendered.append(f"{k}: {v}")
        else:
            rendered.append(f"{k}: {_js_literal(v)}")
    return rendered


def _build_merged_fix_block(event_name, occurrence, issues, spec_index):
    """
    Build one ready-to-paste dataLayer.push block for this event/pattern, merging
    valid fields (echoed from a representative real occurrence), missing fields
    (added as typed placeholders), and wrong-typed fields (replaced with a
    corrected placeholder) into a single multi-line JS block — covering event-scoped
    fields, currency/value/transaction_id and friends, and a representative items[] entry.
    """
    match = spec_index.get(normalize_key(event_name))
    param_spec = match[1].get("params", {}) if match else {}
    item_spec = match[1].get("items", {}) if match else {}

    missing_event_params = {
        i["parameter"] for i in issues if i["kind"] == "missing_param" and i["scope"] == "event"
    }
    type_bad_event_params = {
        i["parameter"] for i in issues if i["kind"] == "type_mismatch" and i["scope"] == "event"
    }
    missing_item_params = {
        i["parameter"] for i in issues if i["kind"] == "missing_param" and i["scope"] == "item"
    }
    type_bad_item_params = {
        i["parameter"] for i in issues if i["kind"] == "type_mismatch" and i["scope"] == "item"
    }
    if any(i["kind"] == "missing_items_array" for i in issues):
        missing_item_params |= {p for p, r in item_spec.items() if r.get("required", True)}

    event_fields = _merge_scope_fields(param_spec, occurrence.get("params", {}), missing_event_params, type_bad_event_params)

    top_entries = [f"event: {json.dumps(event_name)}"]
    top_entries.extend(_render_field_entries(event_fields))

    if item_spec:
        rep_items = occurrence.get("items") or [{}]
        rep_item = rep_items[0] if rep_items else {}
        item_fields = _merge_scope_fields(item_spec, rep_item, missing_item_params, type_bad_item_params)
        item_entry_lines = [f"      {line}" for line in _render_field_entries(item_fields)]
        if item_entry_lines:
            items_block = "items: [\n    {\n" + ",\n".join(item_entry_lines) + "\n    }\n  ]"
            top_entries.append(items_block)

    body = ",\n  ".join(top_entries)
    return f"dataLayer.push({{\n  {body}\n}});"


def _build_event_reports(occurrence_issues, spec_index):
    """
    Pass 3: group occurrences by (event, exact set of issue keys triggered) into
    one aggregated report per distinct breakage pattern, each with every issue for
    that pattern listed together and a single merged fix block.
    """
    groups = OrderedDict()
    for occurrence, canonical_name, issues in occurrence_issues:
        issue_keys = frozenset(issue["key"] for issue in issues)
        group_key = (canonical_name, issue_keys)
        if group_key not in groups:
            groups[group_key] = {
                "event_name": canonical_name,
                "occurrences_affected": 0,
                "issues": issues,
                "representative_occurrence": occurrence,
            }
        groups[group_key]["occurrences_affected"] += 1

    reports = []
    for group in groups.values():
        issues = group["issues"]
        if not issues:
            continue  # a fully clean occurrence pattern — nothing to report
        worst_severity = min((i["severity"] for i in issues), key=lambda t: _TIER_RANK[t])
        fix_block = _build_merged_fix_block(
            group["event_name"], group["representative_occurrence"] or {}, issues, spec_index
        )
        reports.append({
            "event_name": group["event_name"],
            "severity": worst_severity,
            "occurrences_affected": group["occurrences_affected"],
            "issue_count": len(issues),
            "issues": [
                {
                    "severity": i["severity"],
                    "scope": i["scope"],
                    "parameter": i["parameter"],
                    "issue": i["issue"],
                    "expected": i["expected"],
                }
                for i in issues
            ],
            "suggested_fix": fix_block,
        })

    reports.sort(key=lambda r: (_TIER_RANK[r["severity"]], r["event_name"]))
    return reports


def diff_events(observed_events, spec):
    spec_index = build_spec_index(spec)
    total_occurrences = len(observed_events)
    events_seen = set()

    occurrence_issues = []  # list of (occurrence, canonical_event_name, issues) triples
    for occurrence in observed_events:
        events_seen.add(occurrence["event_name"])
        canonical_name, issues = _diff_occurrence(occurrence, spec_index)
        occurrence_issues.append((occurrence, canonical_name, issues))

    findings = _dedup_global_findings(occurrence_issues)
    event_reports = _build_event_reports(occurrence_issues, spec_index)

    # Strip internal-only fields (e.g. _observed_value used for fix-block building)
    # before anything gets serialized — they were never meant to leave this module.
    for f in findings:
        f.pop("_observed_value", None)

    summary = {
        "total_occurrences_analyzed": total_occurrences,
        "distinct_events_seen": len(events_seen),
        "critical_count": sum(1 for f in findings if f["severity"] == CRITICAL),
        "warning_count": sum(1 for f in findings if f["severity"] == WARNING),
        "notice_count": sum(1 for f in findings if f["severity"] == NOTICE),
    }
    return summary, findings, event_reports


# ---------------------------------------------------------------------------
# Optional self-contained Excel export
# ---------------------------------------------------------------------------

_SEVERITY_FILL_COLORS = {
    CRITICAL: "F8D7DA",  # light red
    WARNING: "FFF3CD",   # light amber
    NOTICE: "D1ECF1",    # light blue
}

_EXCEL_HEADERS = ["Severity", "Event", "Occurrences Affected", "Issue", "Suggested Fix"]
_EXCEL_COLUMN_WIDTHS = [11, 18, 12, 60, 55]
_EXCEL_ROW_LINE_HEIGHT = 15  # points per wrapped line, used to reserve enough row height explicitly


def write_excel_findings(event_reports, path):
    """
    Write the aggregated per-event-pattern reports directly to an .xlsx workbook.
    Each individual issue gets its own row (rather than cramming a whole event's
    issues into one newline-joined cell), with the Event / Occurrences Affected /
    Suggested Fix columns merged vertically across that event's rows — this keeps
    every cell's content to a single short block, since some lightweight previewers
    (including in-chat file viewers) don't reliably auto-size row height for wrapped
    multi-line cells, and long wrapped text silently gets clipped in those rather than
    actually wrapping. Row heights are also set explicitly, rather than left to
    auto-fit, for the same reason. This is entirely optional — the JSON output already
    contains both views, and a calling Claude session can always build a nicer-formatted
    workbook itself. This exists for standalone/non-interactive use (e.g. a scheduled
    job in this repo's broader data-quality suite) where nothing downstream converts
    the JSON for you.

    Only imports openpyxl when actually called, so requesting plain JSON output
    never requires installing anything beyond the standard library.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        fail(
            "openpyxl is required to write an Excel findings workbook (--excel-output) but is not installed.",
            "Install it with `pip install openpyxl` (add --break-system-packages if your environment's "
            "Python blocks unmanaged installs), or drop --excel-output and use the JSON output directly — "
            "the JSON already contains everything a workbook would.",
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "GA4 Audit Findings"
    ws.append(_EXCEL_HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.row_dimensions[1].height = 18

    top_left = Alignment(vertical="top", wrap_text=False)
    top_wrap = Alignment(vertical="top", wrap_text=True)

    for report in event_reports:
        issues = report["issues"] or [{
            "severity": report["severity"], "scope": "-", "parameter": None, "issue": "(no individual issues recorded)"
        }]
        fix_text = report.get("suggested_fix", "") or ""
        start_row = ws.max_row + 1

        for idx, issue in enumerate(issues):
            ws.append([
                issue["severity"],
                report["event_name"] if idx == 0 else None,
                report["occurrences_affected"] if idx == 0 else None,
                f"[{issue['scope']}] {issue['parameter'] or '-'}: {issue['issue']}",
                fix_text if idx == 0 else None,
            ])
            row = ws.max_row
            ws.cell(row=row, column=1).alignment = top_left
            ws.cell(row=row, column=2).alignment = top_left
            ws.cell(row=row, column=3).alignment = top_left
            ws.cell(row=row, column=4).alignment = top_wrap
            ws.cell(row=row, column=5).alignment = top_wrap
            fill_color = _SEVERITY_FILL_COLORS.get(issue["severity"])
            if fill_color:
                ws.cell(row=row, column=1).fill = PatternFill(
                    start_color=fill_color, end_color=fill_color, fill_type="solid"
                )
            # Reserve enough height for this row's own Issue text (roughly one line
            # per ~90 chars at the Issue column's width) even before considering the
            # merged Suggested Fix block below.
            issue_lines_needed = max(1, -(-len(issue["issue"]) // 90))
            ws.row_dimensions[row].height = max(
                ws.row_dimensions[row].height or 0, issue_lines_needed * _EXCEL_ROW_LINE_HEIGHT, 15
            )

        end_row = ws.max_row
        if end_row > start_row:
            ws.merge_cells(start_row=start_row, end_row=end_row, start_column=2, end_column=2)
            ws.merge_cells(start_row=start_row, end_row=end_row, start_column=3, end_column=3)
            ws.merge_cells(start_row=start_row, end_row=end_row, start_column=5, end_column=5)

        # Spread the height the merged Suggested Fix block needs across its group's
        # rows, on top of whatever each row already reserved for its own Issue text —
        # set explicitly rather than relying on the previewer to auto-fit a wrapped cell.
        fix_lines_needed = fix_text.count("\n") + 1
        total_rows_in_group = end_row - start_row + 1
        extra_height_per_row = (fix_lines_needed * _EXCEL_ROW_LINE_HEIGHT) / total_rows_in_group
        for r in range(start_row, end_row + 1):
            ws.row_dimensions[r].height = max(ws.row_dimensions[r].height or 0, extra_height_per_row)

    for i, width in enumerate(_EXCEL_COLUMN_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A2"

    try:
        wb.save(path)
    except Exception as e:
        fail(f"Could not write the Excel workbook to '{path}': {e}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Diff an observed GA4 data export against a tracking plan and emit tiered findings as JSON."
    )
    parser.add_argument("--observed", required=True, help="Path to the observed data export.")
    parser.add_argument(
        "--tracking-plan",
        required=False,
        default=None,
        help="Path to a custom tracking plan (.json/.csv/.tsv/.xlsx/.xls). Omit to use the bundled default spec.",
    )
    parser.add_argument(
        "--output",
        required=False,
        default=None,
        help="Path to write JSON findings to. Omit to print to stdout.",
    )
    parser.add_argument(
        "--format",
        required=False,
        default=None,
        choices=["bq_json", "ga4_api_json", "gtm_preview_json", "flat_json", "tabular"],
        help="Force the observed-data format instead of auto-detecting it.",
    )
    parser.add_argument(
        "--excel-output",
        required=False,
        default=None,
        help="Optional path to also write the aggregated event reports as an .xlsx workbook directly "
             "(requires openpyxl). Entirely optional — the JSON output already contains everything needed.",
    )
    args = parser.parse_args()

    try:
        observed_events, detected_format = detect_and_load_observed(args.observed, args.format)

        if args.tracking_plan:
            spec = load_custom_tracking_plan(args.tracking_plan)
            plan_source = f"custom:{args.tracking_plan}"
        else:
            spec = load_default_spec()
            plan_source = "bundled_default"

        summary, findings, event_reports = diff_events(observed_events, spec)
        summary["observed_data_format_detected"] = detected_format
        summary["tracking_plan_source"] = plan_source

        result = {"summary": summary, "findings": findings, "event_reports": event_reports}
        output_text = json.dumps(result, indent=2, default=str)

        if args.output:
            with open(args.output, "w", encoding="utf-8") as f:
                f.write(output_text)
            print(f"Wrote {len(findings)} findings ({len(event_reports)} aggregated event reports) to {args.output}", file=sys.stderr)
        else:
            print(output_text)

        if args.excel_output:
            write_excel_findings(event_reports, args.excel_output)
            print(f"Wrote {len(event_reports)} aggregated event reports to {args.excel_output}", file=sys.stderr)

    except AuditorError as e:
        emit_fatal_and_exit(e)
    except Exception as e:  # last-resort guard against ever hallucinating a "clean" result
        sys.stderr.write(f"UNEXPECTED ERROR: {e}\n")
        sys.stderr.write(
            "HINT: This wasn't one of the anticipated failure modes — re-check the input file by hand "
            "before trusting any partial output above.\n"
        )
        sys.exit(1)


if __name__ == "__main__":
    main()
