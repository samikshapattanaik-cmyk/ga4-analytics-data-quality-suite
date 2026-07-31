#!/usr/bin/env python3
"""
validate_utm.py

Deterministic UTM tag auditor for GA4 campaign URL batches.

Reads a batch of URLs (CSV, XLSX/XLSM, or plain text — one URL per line),
parses utm_source / utm_medium / utm_campaign (and any other utm_* params)
out of each URL, checks them against a taxonomy (a supplied one, or the
built-in GA4-aligned default), and emits structured JSON describing every
issue found, tiered by severity, along with a corrected destination URL
wherever the fix is unambiguous.

Usage:
    python3 validate_utm.py --input urls.csv [--taxonomy taxonomy.json] [--output results.json]

Input formats:
    .csv / .tsv   - must have a column containing URLs (auto-detected by name:
                    'url', 'URL', 'landing page', 'page path', 'campaign url', etc.
                    Falls back to the first column that looks like a URL.)
                    An 'id' / 'ID' / 'name' column is used as the row identifier if present.
    .xlsx / .xlsm - first worksheet, same column auto-detection as CSV.
    .txt          - one URL per line, row ID is the 1-based line number.

Taxonomy JSON format (all keys optional; unspecified keys fall back to defaults):
    {
      "allowed_sources": ["google", "facebook", "newsletter", ...],
      "allowed_mediums": ["cpc", "organic", "email", "social", "paid-social", "referral", "affiliate", "display"],
      "medium_synonyms": {"ppc": "cpc", "e-mail": "email", "paidsocial": "paid-social"},
      "source_synonyms": {"fb": "facebook", "ig": "instagram"},
      "required_params": ["utm_source", "utm_medium", "utm_campaign"]
    }

This script never guesses at data it can't read: any file/parse failure stops
execution with a specific, actionable error message rather than silently
skipping rows or inventing placeholder data.
"""

import argparse
import csv
import json
import re
import sys
from urllib.parse import urlsplit, urlunsplit, parse_qsl, urlencode, unquote

# --------------------------------------------------------------------------
# Built-in GA4-aligned default taxonomy (used whenever --taxonomy is omitted,
# or to fill in any keys the user's taxonomy file doesn't specify).
# Mirrors references/utm-taxonomy-standards.md — keep the two in sync.
# --------------------------------------------------------------------------

DEFAULT_TAXONOMY = {
    "allowed_mediums": [
        "cpc", "organic", "email", "social", "paid-social",
        "referral", "affiliate", "display", "video", "sms", "push", "audio",
    ],
    "medium_synonyms": {
        "ppc": "cpc",
        "cost-per-click": "cpc",
        "costperclick": "cpc",
        "paid": "cpc",
        "paidsearch": "cpc",
        "paid-search": "cpc",
        "e-mail": "email",
        "e_mail": "email",
        "mail": "email",
        "newsletter": "email",
        "paidsocial": "paid-social",
        "paid_social": "paid-social",
        "social-paid": "paid-social",
        "socialpaid": "paid-social",
        "organicsocial": "social",
        "organic-social": "social",
        "organic_social": "social",
        "social-organic": "social",
        "banner": "display",
        "banners": "display",
        "cpm": "display",
        "aff": "affiliate",
        "affiliates": "affiliate",
    },
    "allowed_sources": [],  # empty = don't restrict sources, only check casing/synonyms
    "source_synonyms": {},
    "required_params": ["utm_source", "utm_medium", "utm_campaign"],
}

# Regex patterns used for structural checks -------------------------------

# Unreplaced template placeholders: {var}, {{var}}, %7Bvar%7D (encoded braces),
# <<var>>, [VAR], $var$, ${var}
TEMPLATE_PATTERNS = [
    re.compile(r"\{\{[^{}]+\}\}"),
    re.compile(r"\{[^{}]+\}"),
    re.compile(r"%7[Bb][^%]*%7[Dd]"),
    re.compile(r"<<[^<>]+>>"),
    re.compile(r"\[[A-Za-z0-9_ ]{2,40}\]"),
    re.compile(r"\$\{[^{}]+\}"),
]

# Double-encoding fingerprint: %25 followed by two hex digits means a literal
# "%" was itself percent-encoded, i.e. the string went through encoding twice.
DOUBLE_ENCODING_RE = re.compile(r"%25[0-9A-Fa-f]{2}")

URL_LIKE_HEADER_HINTS = ("url", "link", "landing page", "page path", "campaign url", "destination")
ID_HEADER_HINTS = ("id", "name", "row", "campaign id", "campaign name")


class InputError(Exception):
    """Raised for any problem reading/parsing the input or taxonomy file."""


# --------------------------------------------------------------------------
# File loading
# --------------------------------------------------------------------------

def load_taxonomy(path):
    if path is None:
        return DEFAULT_TAXONOMY
    try:
        with open(path, "r", encoding="utf-8") as f:
            user_taxonomy = json.load(f)
    except FileNotFoundError:
        raise InputError(
            f"Taxonomy file not found: '{path}'. Check the path and try again."
        )
    except json.JSONDecodeError as e:
        raise InputError(
            f"Taxonomy file '{path}' is not valid JSON ({e}). "
            "Check for trailing commas or unquoted keys."
        )
    except Exception as e:
        raise InputError(f"Could not read taxonomy file '{path}': {e}")

    merged = json.loads(json.dumps(DEFAULT_TAXONOMY))  # deep copy
    for key, value in user_taxonomy.items():
        merged[key] = value
    return merged


def _pick_column(header, hints):
    lower = [h.strip().lower() for h in header]
    for hint in hints:
        for i, col in enumerate(lower):
            if hint == col:
                return i
    for hint in hints:
        for i, col in enumerate(lower):
            if hint in col:
                return i
    return None


def _looks_like_url(value):
    return isinstance(value, str) and re.match(r"^https?://", value.strip(), re.IGNORECASE)


def load_rows_from_csv(path):
    try:
        with open(path, "r", encoding="utf-8-sig", newline="") as f:
            sample = f.read(4096)
            f.seek(0)
            try:
                dialect = csv.Sniffer().sniff(sample, delimiters=",\t;")
            except csv.Error:
                dialect = csv.excel
            reader = csv.reader(f, dialect)
            rows = list(reader)
    except FileNotFoundError:
        raise InputError(f"Input file not found: '{path}'. Check the path and try again.")
    except UnicodeDecodeError:
        raise InputError(
            f"'{path}' could not be read as text (encoding issue). "
            "Try re-exporting it as UTF-8 CSV."
        )
    except Exception as e:
        raise InputError(f"Could not read '{path}' as CSV: {e}")

    if not rows:
        raise InputError(f"'{path}' is empty — nothing to audit.")

    header = rows[0]
    url_col = _pick_column(header, URL_LIKE_HEADER_HINTS)
    id_col = _pick_column(header, ID_HEADER_HINTS)

    data_rows = rows[1:]
    if url_col is None:
        # No recognizable header — check if row 0 itself looks like data (no header row at all)
        if _looks_like_url(header[0] if header else ""):
            data_rows = rows
            url_col = 0
            id_col = None
        else:
            raise InputError(
                f"Could not find a URL column in '{path}'. Expected a header like "
                "'URL', 'Landing Page', or 'Campaign URL'. Found columns: "
                f"{header}. Please rename the URL column or specify it manually."
            )

    results = []
    for i, row in enumerate(data_rows, start=1):
        if url_col >= len(row) or not row[url_col].strip():
            continue  # skip genuinely blank rows, don't fabricate a URL for them
        row_id = row[id_col].strip() if (id_col is not None and id_col < len(row) and row[id_col].strip()) else str(i)
        results.append({"id": row_id, "url": row[url_col].strip()})

    if not results:
        raise InputError(f"'{path}' was read successfully but contained no non-empty URLs.")
    return results


def load_rows_from_xlsx(path):
    try:
        import openpyxl
    except ImportError:
        raise InputError(
            "The 'openpyxl' package is required to read .xlsx files but is not installed. "
            "Install it with: pip install openpyxl --break-system-packages"
        )

    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    except FileNotFoundError:
        raise InputError(f"Input file not found: '{path}'. Check the path and try again.")
    except Exception as e:
        raise InputError(
            f"'{path}' could not be opened as an .xlsx file — please check if it is "
            f"formatted correctly, not corrupted, or try re-saving/exporting it. ({e})"
        )

    try:
        ws = wb[wb.sheetnames[0]]
        all_rows = list(ws.iter_rows(values_only=True))
    except Exception as e:
        raise InputError(f"Could not read rows from the first sheet of '{path}': {e}")
    finally:
        wb.close()

    if not all_rows:
        raise InputError(f"'{path}' is empty — nothing to audit.")

    header = [str(c) if c is not None else "" for c in all_rows[0]]
    url_col = _pick_column(header, URL_LIKE_HEADER_HINTS)
    id_col = _pick_column(header, ID_HEADER_HINTS)

    data_rows = all_rows[1:]
    if url_col is None:
        if _looks_like_url(str(header[0]) if header else ""):
            data_rows = all_rows
            url_col = 0
            id_col = None
        else:
            raise InputError(
                f"Could not find a URL column in '{path}'. Expected a header like "
                f"'URL', 'Landing Page', or 'Campaign URL'. Found columns: {header}."
            )

    results = []
    for i, row in enumerate(data_rows, start=1):
        if url_col >= len(row) or row[url_col] is None or not str(row[url_col]).strip():
            continue
        row_id = str(row[id_col]).strip() if (id_col is not None and id_col < len(row) and row[id_col] is not None and str(row[id_col]).strip()) else str(i)
        results.append({"id": row_id, "url": str(row[url_col]).strip()})

    if not results:
        raise InputError(f"'{path}' was read successfully but contained no non-empty URLs.")
    return results


def load_rows_from_txt(path):
    try:
        with open(path, "r", encoding="utf-8") as f:
            lines = [line.strip() for line in f]
    except FileNotFoundError:
        raise InputError(f"Input file not found: '{path}'. Check the path and try again.")
    except UnicodeDecodeError:
        raise InputError(f"'{path}' could not be read as UTF-8 text.")
    except Exception as e:
        raise InputError(f"Could not read '{path}': {e}")

    results = [{"id": str(i), "url": line} for i, line in enumerate(lines, start=1) if line]
    if not results:
        raise InputError(f"'{path}' is empty — nothing to audit.")
    return results


def load_rows(path):
    lower = path.lower()
    if lower.endswith(".csv") or lower.endswith(".tsv"):
        return load_rows_from_csv(path)
    elif lower.endswith(".xlsx") or lower.endswith(".xlsm"):
        return load_rows_from_xlsx(path)
    elif lower.endswith(".txt"):
        return load_rows_from_txt(path)
    else:
        raise InputError(
            f"Unrecognized file type for '{path}'. Supported: .csv, .tsv, .xlsx, .xlsm, .txt "
            "(one URL per line). Please re-export your data in one of these formats."
        )


# --------------------------------------------------------------------------
# URL parsing / auditing
# --------------------------------------------------------------------------

def find_template_placeholder(value):
    if not value:
        return None
    for pattern in TEMPLATE_PATTERNS:
        m = pattern.search(value)
        if m:
            return m.group(0)
    return None


def fully_unquote(value, max_passes=3):
    """Repeatedly unquote until stable or max_passes reached, tracking pass count."""
    current = value
    passes = 0
    for _ in range(max_passes):
        nxt = unquote(current)
        if nxt == current:
            break
        current = nxt
        passes += 1
    return current, passes


def audit_url(row_id, raw_url, taxonomy):
    issues = []
    parsed = {"utm_source": None, "utm_medium": None, "utm_campaign": None}
    corrected_url = raw_url

    # --- Structural validity check ---
    try:
        parts = urlsplit(raw_url)
        if not parts.scheme or not parts.netloc:
            issues.append({
                "tier": "CRITICAL",
                "code": "malformed_url",
                "message": f"URL is missing a scheme or domain and cannot be parsed as valid: '{raw_url}'",
            })
            return {
                "id": row_id, "original_url": raw_url, "issues": issues,
                "corrected_url": None, "parsed": parsed,
            }
    except Exception as e:
        issues.append({
            "tier": "CRITICAL",
            "code": "malformed_url",
            "message": f"URL could not be parsed at all ({e}): '{raw_url}'",
        })
        return {
            "id": row_id, "original_url": raw_url, "issues": issues,
            "corrected_url": None, "parsed": parsed,
        }

    # Parse manually (rather than via parse_qsl, which decodes eagerly) so we can
    # inspect the *raw* still-encoded value first and reliably detect double-encoding
    # before any decoding happens.
    raw_pairs = []
    if parts.query:
        for chunk in parts.query.split("&"):
            if not chunk:
                continue
            if "=" in chunk:
                raw_key, raw_value = chunk.split("=", 1)
            else:
                raw_key, raw_value = chunk, ""
            raw_pairs.append((unquote(raw_key), raw_value))

    query_dict = {}
    corrected_pairs = []

    for key, raw_value in raw_pairs:
        original_value = raw_value

        # Double-encoding check on the RAW (still percent-encoded) value.
        if DOUBLE_ENCODING_RE.search(raw_value):
            decoded_once = unquote(raw_value)
            decoded_fully, _ = fully_unquote(raw_value, max_passes=3)
            issues.append({
                "tier": "WARNING",
                "code": "double_encoding",
                "message": f"'{key}' is double-encoded ('{original_value}'); decodes to '{decoded_once}' "
                           f"once, or '{decoded_fully}' fully.",
            })
            value = decoded_fully
        else:
            value = unquote(raw_value)

        # Template placeholder check
        placeholder = find_template_placeholder(value) or find_template_placeholder(original_value)
        if placeholder:
            issues.append({
                "tier": "CRITICAL",
                "code": "unfilled_template_variable",
                "message": f"'{key}' still contains an unreplaced template placeholder: '{placeholder}'. "
                           "This param must be filled in manually — no automatic fix is possible.",
            })

        query_dict[key] = value

        if key in ("utm_source", "utm_medium", "utm_campaign"):
            parsed[key] = value

        corrected_pairs.append([key, value, placeholder is not None])

    # --- Required param checks ---
    required = taxonomy.get("required_params", DEFAULT_TAXONOMY["required_params"])
    for req in required:
        if req not in query_dict or not query_dict[req]:
            tier = "CRITICAL" if req == "utm_source" else "WARNING"
            issues.append({
                "tier": tier,
                "code": f"missing_{req}",
                "message": f"'{req}' is missing entirely from this URL.",
            })

    # --- Medium synonym / allow-list check ---
    medium_synonyms = taxonomy.get("medium_synonyms", {})
    allowed_mediums = set(taxonomy.get("allowed_mediums", []))
    if parsed["utm_medium"]:
        medium_raw = parsed["utm_medium"]
        medium_lower = medium_raw.strip().lower()
        canonical_medium = medium_synonyms.get(medium_lower, medium_lower)
        if medium_raw != medium_raw.lower():
            issues.append({
                "tier": "NOTICE",
                "code": "medium_casing",
                "message": f"utm_medium '{medium_raw}' has inconsistent casing; GA4 will treat it as "
                           f"distinct from '{medium_raw.lower()}'.",
            })
        if medium_lower in medium_synonyms:
            issues.append({
                "tier": "WARNING",
                "code": "medium_synonym_collision",
                "message": f"utm_medium '{medium_raw}' is a non-standard synonym for "
                           f"'{canonical_medium}' and may not match GA4's channel grouping rules.",
            })
        elif allowed_mediums and canonical_medium not in allowed_mediums:
            issues.append({
                "tier": "WARNING",
                "code": "medium_not_in_taxonomy",
                "message": f"utm_medium '{medium_raw}' is not in the approved taxonomy list.",
            })
        for key, value, _ in corrected_pairs:
            pass
        # write correction back
        for pair in corrected_pairs:
            if pair[0] == "utm_medium":
                pair[1] = canonical_medium

    # --- Source synonym / casing check ---
    source_synonyms = taxonomy.get("source_synonyms", {})
    allowed_sources = set(taxonomy.get("allowed_sources", []))
    if parsed["utm_source"]:
        source_raw = parsed["utm_source"]
        source_lower = source_raw.strip().lower()
        canonical_source = source_synonyms.get(source_lower, source_lower)
        if source_raw != source_raw.lower():
            issues.append({
                "tier": "NOTICE",
                "code": "source_casing",
                "message": f"utm_source '{source_raw}' has inconsistent casing; GA4 will treat it as "
                           f"distinct from '{source_raw.lower()}'.",
            })
        if source_lower in source_synonyms:
            issues.append({
                "tier": "WARNING",
                "code": "source_synonym_collision",
                "message": f"utm_source '{source_raw}' is a non-standard synonym for '{canonical_source}'.",
            })
        elif allowed_sources and canonical_source not in allowed_sources:
            issues.append({
                "tier": "WARNING",
                "code": "source_not_in_taxonomy",
                "message": f"utm_source '{source_raw}' is not in the approved source list.",
            })
        for pair in corrected_pairs:
            if pair[0] == "utm_source":
                pair[1] = canonical_source

    # --- Campaign casing (notice only; campaign values are freer-form) ---
    if parsed["utm_campaign"]:
        campaign_raw = parsed["utm_campaign"]
        if campaign_raw != campaign_raw.lower():
            issues.append({
                "tier": "NOTICE",
                "code": "campaign_casing",
                "message": f"utm_campaign '{campaign_raw}' has inconsistent casing, which can create "
                           "duplicate campaign rows in reports.",
            })
        for pair in corrected_pairs:
            if pair[0] == "utm_campaign":
                pair[1] = campaign_raw.lower()

    # --- Build corrected URL (skip if any param still has an unfilled placeholder) ---
    has_unfillable = any(has_placeholder for _, _, has_placeholder in corrected_pairs)
    if has_unfillable:
        corrected_url = None
    else:
        new_query = urlencode([(k, v) for k, v, _ in corrected_pairs])
        corrected_url = urlunsplit((parts.scheme, parts.netloc, parts.path, new_query, parts.fragment))

    return {
        "id": row_id,
        "original_url": raw_url,
        "issues": issues,
        "corrected_url": corrected_url,
        "parsed": parsed,
    }


def worst_tier(issues):
    if any(i["tier"] == "CRITICAL" for i in issues):
        return "CRITICAL"
    if any(i["tier"] == "WARNING" for i in issues):
        return "WARNING"
    if any(i["tier"] == "NOTICE" for i in issues):
        return "NOTICE"
    return "OK"


# --------------------------------------------------------------------------
# Main
# --------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Audit a batch of campaign URLs for UTM tagging issues.")
    parser.add_argument("--input", required=True, help="Path to input file (.csv, .tsv, .xlsx, .xlsm, or .txt)")
    parser.add_argument("--taxonomy", default=None, help="Path to a taxonomy JSON file (optional; defaults to built-in GA4 baseline)")
    parser.add_argument("--output", default=None, help="Path to write JSON results to (defaults to stdout)")
    args = parser.parse_args()

    try:
        taxonomy = load_taxonomy(args.taxonomy)
        rows = load_rows(args.input)
    except InputError as e:
        sys.stderr.write(f"ERROR: {e}\n")
        sys.exit(1)

    results = [audit_url(row["id"], row["url"], taxonomy) for row in rows]

    summary = {"total": len(results), "critical": 0, "warning": 0, "notice": 0, "ok": 0}
    for r in results:
        tier = worst_tier(r["issues"])
        r["worst_tier"] = tier
        summary[tier.lower()] += 1

    output = {"summary": summary, "results": results}
    output_json = json.dumps(output, indent=2)

    if args.output:
        try:
            with open(args.output, "w", encoding="utf-8") as f:
                f.write(output_json)
        except Exception as e:
            sys.stderr.write(f"ERROR: Could not write output to '{args.output}': {e}\n")
            sys.exit(1)
        print(f"Wrote results for {summary['total']} URLs to {args.output} "
              f"({summary['critical']} critical, {summary['warning']} warning, "
              f"{summary['notice']} notice, {summary['ok']} clean).")
    else:
        print(output_json)


if __name__ == "__main__":
    main()
